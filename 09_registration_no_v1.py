from tkinter import *
from tkinter import Toplevel
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl  # , xlrd
from openpyxl import Workbook
import pathlib

# modules required:
# pip install pathlib
# pip install openpyxl
# pip install xlrd
# pip install pillow

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("Fossil Registration System")
root.geometry("1250x740+210+100")
root.config(bg=background)

file = pathlib.Path('Fossil_Data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration Number"
    sheet['B1'] = "Catalogue Number"
    sheet['C1'] = "Species"
    sheet['D1'] = "Date Discovered"
    sheet['E1'] = "Date Registered"
    sheet['F1'] = "Condition"
    sheet['G1'] = "Diet"
    sheet['H1'] = "Nickname"
    sheet['I1'] = "Estimated Period"
    sheet['J1'] = "Taxonomy"
    sheet['K1'] = "Additional Details"
    sheet['L1'] = "Traits"

    file.save('Fossil_Data.xlsx')

# ############ Registration Number
# automatically assigns new save to entry system
def registration_no():
    file = openpyxl.load_workbook('Fossil_Data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)

    except:
        Registration.set(1)  # was originally "1" but had warning - didn't break anything

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=200, y=150)
reg_entry.config(state='readonly')

registration_no()


root.mainloop()
