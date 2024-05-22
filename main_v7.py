from tkinter import *
from tkinter import Toplevel
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
import pathlib

background = "#0D3B29"
framebg = "#EDEDED"
framefg = "#06283D"
buttonbg = "#1F5327"

root = Tk()
root.title("Fossil Registration System")
root.geometry("960x1000+210+100")
root.config(bg=background)

file = pathlib.Path('Fossil_Data.xlsx')
if file.exists():
    pass
else:
    file = openpyxl.Workbook()
    sheet = file.active
    sheet['A1'] = "Registration Number"
    sheet['B1'] = "Catalogue Number"
    sheet['C1'] = "Species"
    sheet['D1'] = "Date Discovered"
    sheet['E1'] = "Date Registered"
    sheet['F1'] = "Condition"
    sheet['G1'] = "Diet"
    sheet['H1'] = "Nickname"
    sheet['I1'] = "Estimated Period"
    sheet['J1'] = "Taxonomy"
    sheet['K1'] = "Additional Details"
    sheet['L1'] = "Traits"

    file.save('Fossil_Data.xlsx')


def Exit():
    root.destroy()


def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select Image File") # filetype=(("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All files", "*.txt")))

    img = Image.open(filename)
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


def registration_no():
    file = openpyxl.load_workbook('Fossil_Data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set(1)


def Clear():
    global img
    CatNum.set('')
    DDate.set('')
    Species.set('')
    Nickname.set('')
    EstimatedP.set('')
    Taxonomy.set('')
    Details.set('')
    Traits.set('')
    Condition.set("Select Condition")

    registration_no()

    saveButton.config(state='normal')
    cat_entry.config(state='normal')

    img1 = PhotoImage(file='Images/no_img.png')
    lbl.config(image=img1)
    lbl.image = img1

    img = ""

    radio.set(0)


def Save():
    R1 = Registration.get()
    CN1 = CatNum.get()
    S1 = Species.get()
    D2 = DDate.get()
    D3 = Date.get()
    C1 = Condition.get()
    D1 = Diet
    N1 = Nickname.get()
    EP1 = EstimatedP.get()
    T1 = Taxonomy.get()
    D4 = Details.get()
    T2 = Traits.get()

    if CN1 == "" or C1 == "Select Condition" or D2 == "" or S1 == "" \
            or N1 == "" or EP1 == "" or T1 == "" or D4 == "" or T2 == "":
        messagebox.showerror("Error", "All Data Must Be Entered!")
    else:
        file = openpyxl.load_workbook('Fossil_Data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=CN1)
        sheet.cell(column=3, row=sheet.max_row, value=S1)
        sheet.cell(column=4, row=sheet.max_row, value=D2)
        sheet.cell(column=5, row=sheet.max_row, value=D3)
        sheet.cell(column=6, row=sheet.max_row, value=C1)
        sheet.cell(column=7, row=sheet.max_row, value=D1)
        sheet.cell(column=8, row=sheet.max_row, value=N1)
        sheet.cell(column=9, row=sheet.max_row, value=EP1)
        sheet.cell(column=10, row=sheet.max_row, value=T1)
        sheet.cell(column=11, row=sheet.max_row, value=D4)
        sheet.cell(column=12, row=sheet.max_row, value=T2)

        file.save(r'Fossil_Data.xlsx')

        try:
            img.save("Fossil Images/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo("Warning", "No Image Assigned")

        messagebox.showinfo("Info", "Data Input Successful")

        Clear()

        registration_no()


def search():
    text = Search.get()

    Clear()
    saveButton.config(state='disabled')
    cat_entry.config(state='disabled', bg='white', fg='black')

    file = openpyxl.load_workbook("Fossil_Data.xlsx")
    sheet = file.active

    try:
        for row in sheet.rows:
            if row[0].value == int(text):
                catalogue_no = row[0]
                reg_no_position = str(catalogue_no)[14:-1]
                reg_number = str(catalogue_no)[15:-1]

        x1 = sheet.cell(row=int(reg_number), column=1).value
        x2 = sheet.cell(row=int(reg_number), column=2).value
        x3 = sheet.cell(row=int(reg_number), column=3).value
        x4 = sheet.cell(row=int(reg_number), column=4).value
        x5 = sheet.cell(row=int(reg_number), column=5).value
        x6 = sheet.cell(row=int(reg_number), column=6).value
        x7 = sheet.cell(row=int(reg_number), column=7).value
        x8 = sheet.cell(row=int(reg_number), column=8).value
        x9 = sheet.cell(row=int(reg_number), column=9).value
        x10 = sheet.cell(row=int(reg_number), column=10).value
        x11 = sheet.cell(row=int(reg_number), column=11).value
        x12 = sheet.cell(row=int(reg_number), column=12).value

        Registration.set(x1)
        CatNum.set(x2)
        Species.set(x3)
        DDate.set(x4)
        Date.set(x5)
        Condition.set(x6)

        if x7 == 'Herbivore':
            R1.select()
        elif x7 == 'Carnivore':
            R2.select()
        elif x7 == 'Omnivore':
            R3.select()
        else:
            R4.select()

        Nickname.set(x8)
        EstimatedP.set(x9)
        Taxonomy.set(x10)
        Details.set(x11)
        Traits.set(x12)

        try:
            img = (Image.open("Fossil Images/" + str(x2) + ".jpg"))
            resized_image = img.resize((190, 190))
            photo2 = ImageTk.PhotoImage(resized_image)
            lbl.config(image=photo2)
            lbl.image = photo2
        except FileNotFoundError:
            pass

    except:
        messagebox.showerror("Invalid", "Invalid Registration Number")
        Clear()


def Update():
    R1 = Registration.get()
    CN1 = CatNum.get()
    S1 = Species.get()
    D2 = DDate.get()
    D3 = Date.get()
    C1 = Condition.get()
    selection()
    D1 = Diet
    N1 = Nickname.get()
    EP1 = EstimatedP.get()
    T1 = Taxonomy.get()
    D4 = Details.get()
    T2 = Traits.get()

    file = openpyxl.load_workbook("Fossil_Data.xlsx")
    sheet = file.active

    reg_number = None
    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]

    if reg_number is not None:
        sheet.cell(column=1, row=int(reg_number), value=R1)
        sheet.cell(column=2, row=int(reg_number), value=CN1)
        sheet.cell(column=3, row=int(reg_number), value=S1)
        sheet.cell(column=4, row=int(reg_number), value=D2)
        sheet.cell(column=5, row=int(reg_number), value=D3)
        sheet.cell(column=6, row=int(reg_number), value=C1)
        sheet.cell(column=7, row=int(reg_number), value=D1)
        sheet.cell(column=8, row=int(reg_number), value=N1)
        sheet.cell(column=9, row=int(reg_number), value=EP1)
        sheet.cell(column=10, row=int(reg_number), value=T1)
        sheet.cell(column=11, row=int(reg_number), value=D4)
        sheet.cell(column=12, row=int(reg_number), value=T2)

        file.save(r"Fossil_Data.xlsx")

        try:
            img.save("Fossil Images/" + str(R1) + ".jpg")
            messagebox.showinfo("Info", "Data Input Successful")
        except:
            messagebox.showinfo("Warning", "No Image Assigned, Data Input Successful")

        Clear()  # clear selection
    else:
        messagebox.showerror("Error", "Registration Number not found!")


def display_help():
    help_window = Toplevel(root)
    help_window.title("Help")
    help_window.config(bg=background)

    help_info = "The Fossil Registration System is a user-friendly application designed to " \
                "manage and track information about fossils. To begin, users can either search " \
                "for existing fossil records using the 'Search' button by entering the registration " \
                "number or start fresh by filling in the details of a new fossil. When entering " \
                "data for a new fossil, users need to provide information such as the catalogue number," \
                " species, discovery date, dietary classification, condition, nickname, estimated period," \
                " taxonomy, and additional details. Users can also upload an image of the fossil. " \
                "The 'Save' button stores the entered data in an Excel file, and the 'Update File' " \
                "button allows users to modify existing records. The 'Clear' button resets the input " \
                "fields for entering new data. For assistance, users can click the 'Help' button, " \
                "which opens a window with helpful information. " \
                "" \
                "Avoid using special characters for the catalogue number as this may cause errors with the images"

    heading = Label(help_window, text="Help", font="arial 20 bold", bg=background, fg="white")
    heading.pack(pady=10)

    wrap_length = 500
    font_size = 12

    help_text = Label(help_window, text=help_info,
                      font=f"arial {font_size}", bg=background, fg="white", wraplength=wrap_length)
    help_text.pack(padx=20, pady=20)

    text_height = help_text.winfo_reqheight() + heading.winfo_reqheight() + 40

    help_window.geometry(f"{wrap_length+40}x{text_height+40}")


def selection():
    global Diet
    Diet = "Unknown"
    value = radio.get()
    if value == 1:
        Diet = "Herbivore"
    elif value == 2:
        Diet = "Carnivore"
    elif value == 3:
        Diet = "Omnivore"
    else:
        Diet = "Unknown"


Label(root, text="Diego Becker da Beast", width=10, height=3,
      bg="#2D6E44", anchor='e').pack(side=TOP, fill=X)

Label(root, text="Fossil Registration", width=10, height=2,
      bg="#1F5327", fg='#fff', font='arial 20 bold').pack(side=TOP, fill=X)

Search = StringVar()
Entry(root, textvariable=Search, width=10, bd=2, font="arial 20").place(x=630, y=70)

Srch = Button(root, text="Search", compound=LEFT, width=12, pady=6,
              bg=buttonbg, fg="white", font="arial 13 bold", command=search)
Srch.place(x=798, y=66)

update_button = Button(root, text="Update File", bg=buttonbg,
                       font="arial 13 bold", fg="white", pady=6, width=10, command=Update)
update_button.place(x=30, y=64)

help_button = Button(root, text="Help", bg=buttonbg, font="arial 13 bold", fg="white", pady=6, width=10, command=display_help)
help_button.place(x=180, y=64)

Registration = IntVar()
Date = StringVar()

Label(root, text="Registration Number:", font="arial 13", bg=background, fg="#fff").place(x=30, y=150)
reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=200, y=150)
reg_entry.config(state='readonly')

registration_no()

Label(root, text="Date Created:", font="arial 13", bg=background, fg="#fff").place(x=30, y=250)
today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=200, y=250)
date_entry.config(state='readonly')

Date.set(d1)

obj = LabelFrame(root, text="Fossil Details", font=20, bd=2, width=900,
                 bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=360)

Label(obj, text="Catalogue Number:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Discovery Date:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Dietary Classification:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text="Species:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Condition:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Nickname:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

CatNum = StringVar()
cat_entry = Entry(obj, textvariable=CatNum, width=20, font="arial 10")
cat_entry.place(x=200, y=50)

DDate = StringVar()
ddate_entry = Entry(obj, textvariable=DDate, width=20, font="arial 10")
ddate_entry.place(x=200, y=100)

radio = IntVar()
R1 = Radiobutton(obj, text="Herbivore", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=20, y=180)

R2 = Radiobutton(obj, text="Carnivore", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=120, y=180)

R3 = Radiobutton(obj, text="Omnivore", variable=radio, value=3, bg=framebg, fg=framefg, command=selection)
R3.place(x=220, y=180)

R4 = Radiobutton(obj, text="Unknown", variable=radio, value=4, bg=framebg, fg=framefg, command=selection)
R4.place(x=320, y=180)

Condition = Combobox(obj, values=['Excellent', 'Good', 'Fair', 'Poor', 'Bad', 'Ruined'],
                     font="Roboto", width=17, state="r")
Condition.place(x=630, y=100)
Condition.set("Select Condition")

Species = StringVar()
species_entry = Entry(obj, textvariable=Species, width=20, font="arial 10")
species_entry.place(x=630, y=50)

Nickname = StringVar()
nickname_entry = Entry(obj, textvariable=Nickname, width=20, font="arial 10")
nickname_entry.place(x=630, y=150)

obj2 = LabelFrame(root, text="Additional Details:", font=20, bd=2, width=900,
                  bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj2.place(x=30, y=640)

Label(obj2, text="Estimated Period:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
EstimatedP = StringVar()
ep_entry = Entry(obj2, textvariable=EstimatedP, width=20, font="arial 10")
ep_entry.place(x=200, y=50)

Label(obj2, text="Details of Fossil Condition:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Details = StringVar()
details_entry = Entry(obj2, textvariable=Details, width=50, font="arial 10")
details_entry.place(x=30, y=150)

Label(obj2, text="Genus:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Taxonomy = StringVar()
tax_entry = Entry(obj2, textvariable=Taxonomy, width=20, font="arial 10")
tax_entry.place(x=630, y=50)

Label(obj2, text="Distinguishable Phenotypes:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Traits = StringVar()
traits_entry = Entry(obj2, textvariable=Traits, width=50, font="arial 10")
traits_entry.place(x=500, y=150)

f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=500, y=130)

img = PhotoImage(file="Images/no_img.png")
lbl = Label(f, bg="black", image=img, width=200, height=200)
lbl.place(x=0, y=0)

Button(root, text="Upload", width=19, height=2, font="arial 12 bold",
       bg="#8FBC8F", command=showimage).place(x=30, y=930)

saveButton = Button(root, text="Save", width=19, height=2,
                    font="arial 12 bold", bg="#8FBC8F", command=Save)
saveButton.place(x=265, y=930)

Button(root, text="Reset", width=19, height=2, font="arial 12 bold",
       bg="#8FBC8F", command=Clear).place(x=500, y=930)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold",
       bg="#8FBC8F", command=Exit).place(x=735, y=930)

root.mainloop()
