from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl  # , xlrd
from openpyxl import Workbook
import pathlib

# modules required:
# pip install pathlib
# pip install openpyxl
# pip install xlrd
# pip install pillow

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("Fossil Registration System")
root.geometry("1250x740+210+100")
root.config(bg=background)

file = pathlib.Path('Fossil_Data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration Number"
    sheet['B1'] = "Catalogue Number"
    sheet['C1'] = "Species"
    sheet['D1'] = "Date Discovered"
    sheet['E1'] = "Date Registered"
    sheet['F1'] = "Condition"
    sheet['G1'] = "Diet"
    sheet['H1'] = "Nickname"
    sheet['I1'] = "Estimated Period"
    sheet['J1'] = "Taxonomy"
    sheet['K1'] = "Additional Details"
    sheet['L1'] = "Traits"

    file.save('Fossil_Data.xlsx')


# ################ exit button, closes window
def Exit():
    root.destroy()


# ################ image upload function, opens file window and replaces image
def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select Image File", filetype=(("JPG File", "*.jpg"),("PNG File", "*.png"),("All files", "*.txt"))) # this code was origionally contained in the image file bit
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


# ############ Registration Number
# automatically assigns new save to entry system
def registration_no():
    file = openpyxl.load_workbook('Fossil_Data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)

    except:
        Registration.set(1)  # was originally "1" but had warning - didn't break anything


# ################ Clear
def Clear():
    global img
    CatNum.set('')
    DDate.set('')
    Species.set('')
    Nickname.set('')
    EstimatedP.set('')
    Taxonomy.set('')
    Details.set('')
    Traits.set('')
    Condition.set("Select Condition")

    registration_no()

    saveButton.config(state='normal')
    cat_entry.config(state='normal')

    img1 = PhotoImage(file='Images/no_img.png')
    lbl.config(image=img1)
    lbl.image = img1

    img = ""


# ########## Save function
def Save():
    R1 = Registration.get()
    CN1 = CatNum.get()
    S1 = Species.get()
    D2 = DDate.get()
    D3 = Date.get()
    C1 = Condition.get()  # make so it does print "select condition"
    D1 = Diet
    N1 = Nickname.get()
    EP1 = EstimatedP.get()
    T1 = Taxonomy.get()
    D4 = Details.get()
    T2 = Traits.get()

    if CN1 == "" or C1 == "Select Condition" or D2 == "" or S1 == "" \
            or N1 == "" or EP1 == "" or T1 == "" or D4 == "" or T2 == "":
        messagebox.showerror("Error", "All Data Must Be Entered!")
    else:
        file = openpyxl.load_workbook('Fossil_Data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=CN1)
        sheet.cell(column=3, row=sheet.max_row, value=S1)
        sheet.cell(column=4, row=sheet.max_row, value=D2)
        sheet.cell(column=5, row=sheet.max_row, value=D3)
        sheet.cell(column=6, row=sheet.max_row, value=C1)
        sheet.cell(column=7, row=sheet.max_row, value=D1)
        sheet.cell(column=8, row=sheet.max_row, value=N1)
        sheet.cell(column=9, row=sheet.max_row, value=EP1)
        sheet.cell(column=10, row=sheet.max_row, value=T1)
        sheet.cell(column=11, row=sheet.max_row, value=D4)
        sheet.cell(column=12, row=sheet.max_row, value=T2)

        file.save(r'Fossil_Data.xlsx')

        try:
            img.save("Fossil Images/" + str(CN1) + ".jpg")
            messagebox.showinfo("Info", "Data Input Successful")
        except:
            messagebox.showinfo("Warning", "No Image Assigned, Data Input Successful")

        Clear()  # clear selection

        registration_no()  # recheck registration number


# ############## Search Function
def search():
    text = Search.get()  # taking input from entry box

    Clear()  # to clear current data when something else is searched
    saveButton.config(state='disabled')  # after clicking on search, save button disables so that nobody can click on it
    cat_entry.config(state='disabled', bg='white', fg='black')  # after clicking on search, CN label disables

    file = openpyxl.load_workbook("Fossil_Data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            catalogue_no = row[0]
            # print(str(catalogue_no))
            reg_no_position = str(catalogue_no)[14:-1]
            reg_number = str(catalogue_no)[15:-1]

            # print(reg_no_position)
            # print(reg_number)
    try:
        # print(str(catalogue_no))
        pass
    except:
        messagebox.showerror("Invalid", "Invalid Registration Number")

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value

    # print(x1)
    # print(x2)
    # print(x3)
    # print(x4)
    # print(x5)
    # print(x6)
    # print(x7)
    # print(x8)
    # print(x9)
    # print(x10)
    # print(x11)
    # print(x12)

    Registration.set(x1)
    CatNum.set(x2)
    Species.set(x3)
    DDate.set(x4)
    Date.set(x5)
    Condition.set(x6)

    if x7 == 'Herbivore':
        R1.select()
    elif x7 == 'Carnivore':
        R2.select()
    elif x7 == 'Omnivore':
        R3.select()
    else:
        R4.select()

    Nickname.set(x8)
    EstimatedP.set(x9)
    Taxonomy.set(x10)
    Details.set(x11)
    Traits.set(x12)

    img = (Image.open("Fossil Images/" + str(x2) + ".jpg"))  # finds image
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


# ############ Update file
def Update():
    R1 = Registration.get()
    CN1 = CatNum.get()
    S1 = Species.get()
    D2 = DDate.get()
    D3 = Date.get()
    C1 = Condition.get()
    selection()
    D1 = Diet
    N1 = Nickname.get()
    EP1 = EstimatedP.get()
    T1 = Taxonomy.get()
    D4 = Details.get()
    T2 = Traits.get()

    file = openpyxl.load_workbook("Fossil_Data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]

    sheet.cell(column=1, row=int(reg_number), value=R1)
    sheet.cell(column=2, row=int(reg_number), value=CN1)
    sheet.cell(column=3, row=int(reg_number), value=S1)
    sheet.cell(column=4, row=int(reg_number), value=D2)
    sheet.cell(column=5, row=int(reg_number), value=D3)
    sheet.cell(column=6, row=int(reg_number), value=C1)
    sheet.cell(column=7, row=int(reg_number), value=D1)
    sheet.cell(column=8, row=int(reg_number), value=N1)
    sheet.cell(column=9, row=int(reg_number), value=EP1)
    sheet.cell(column=10, row=int(reg_number), value=T1)
    sheet.cell(column=11, row=int(reg_number), value=D4)
    sheet.cell(column=12, row=int(reg_number), value=T2)

    file.save(r"Fossil_Data.xlsx")

    try:
        img.save("Fossil Images/" + str(CN1) + ".jpg")
    except:
        pass

    messagebox.showinfo("Update Status:", "Update Successful")

    Clear()


# ############ dietary classification checkbox
def selection():
    global Diet
    Diet = "Unknown"
    value = radio.get()
    if value == 1:
        Diet = "Herbivore"
    elif value == 2:
        Diet = "Carnivore"
    elif value == 3:
        Diet = "Omnivore"
    else:
        Diet = "Unknown"

    # print(Diet)  # testing purposes


# top frames
Label(root, text="Diego Becker da Beast", width=10, height=3,
      bg="#f0687c", anchor='e').pack(side=TOP, fill=X)

Label(root, text="Fossil Registration", width=10, height=2,
      bg="#c36464", fg='#fff', font='arial 20 bold').pack(side=TOP, fill=X)

# search box to update
# not that Search and search are 2 different variables
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=70)

# image_icon3 = PhotoImage(file="Images/ant.png")
Srch = Button(root, text="Search", compound=LEFT, width=12, pady=6,
              bg='#68ddfa', font="arial 13 bold", command=search)
Srch.place(x=1060, y=66)

# image_icon4 = PhotoImage(fiole="Images/Layer 4.png")
# update button for all details
update_button = Button(root, text="Update File", bg="#68ddfa",
                       font="arial 13 bold", pady=6, width=10, command=Update)
update_button.place(x=110, y=64)
update_button = Button(root, text="Help", bg="#68ddfa",
                       font="arial 13 bold", pady=6, width=10)
update_button.place(x=250, y=64)

# Registration number and date - same row
Label(root, text="Registration Number:", font="arial 13", fg=framebg,
      bg=background).place(x=30, y=150)
Label(root, text="Date:", font="arial 13", fg=framebg,
      bg=background).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=200, y=150)
reg_entry.config(state='readonly')

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)
date_entry.config(state='readonly')

Date.set(d1)

# Fossil details
###########
obj = LabelFrame(root, text="Fossil Details", font=20, bd=2, width=900,
                 bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Catalogue Number:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Discovery Date:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Dietary Classification:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text="Species:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Condition:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Nickname", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

# input box for catalogue number
CatNum = StringVar()
cat_entry = Entry(obj, textvariable=CatNum, width=20, font="arial 10")
cat_entry.place(x=200, y=50)

# input box for date discovered -
DDate = StringVar()
ddate_entry = Entry(obj, textvariable=DDate, width=20, font="arial 10")
ddate_entry.place(x=200, y=100)

# buttons for diet
radio = IntVar()
R1 = Radiobutton(obj, text="Herbivore", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=20, y=180)

R2 = Radiobutton(obj, text="Carnivore", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=120, y=180)

R3 = Radiobutton(obj, text="Omnivore", variable=radio, value=3, bg=framebg, fg=framefg, command=selection)
R3.place(x=220, y=180)

R4 = Radiobutton(obj, text="Unknown", variable=radio, value=4, bg=framebg, fg=framefg, command=selection)
R4.place(x=320, y=180)

# dropdown for condition
Condition = Combobox(obj, values=['Excellent', 'Good', 'Fair', 'Poor', 'Bad', 'Ruined'],
                     font="Roboto", width=17, state="r")
Condition.place(x=630, y=100)
Condition.set("Select Condition")

# input box for species
Species = StringVar()
species_entry = Entry(obj, textvariable=Species, width=20, font="arial 10")
species_entry.place(x=630, y=50)

# input box for nickname
Nickname = StringVar()
nickname_entry = Entry(obj, textvariable=Nickname, width=20, font="arial 10")
nickname_entry.place(x=630, y=150)

#####

# additional details
#########
obj2 = LabelFrame(root, text="Additional Details", font=20, bd=2, width=900,
                  bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Estimated Period:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
EstimatedP = StringVar()
ep_entry = Entry(obj2, textvariable=EstimatedP, width=20, font="arial 10")
ep_entry.place(x=200, y=50)

Label(obj2, text="Details of Fossil Condition:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Details = StringVar()
details_entry = Entry(obj2, textvariable=Details, width=50, font="arial 10")
details_entry.place(x=30, y=150)

Label(obj2, text="Genus:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Taxonomy = StringVar()
tax_entry = Entry(obj2, textvariable=Taxonomy, width=20, font="arial 10")
tax_entry.place(x=630, y=50)

Label(obj2, text="Distinguishable Phenotypes:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Traits = StringVar()
traits_entry = Entry(obj2, textvariable=Traits, width=50, font="arial 10")
traits_entry.place(x=500, y=150)
######

# image
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="Images/no_img.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# buttons for image
Button(root, text="Upload", width=19, height=2, font="arial 12 bold",
       bg="lightblue", command=showimage).place(x=1000, y=370)

saveButton = Button(root, text="Save", width=19, height=2,
                    font="arial 12 bold", bg="lightgreen", command=Save)
saveButton.place(x=1000, y=450)

Button(root, text="Reset", width=19, height=2, font="arial 12 bold",
       bg="lightpink", command=Clear).place(x=1000, y=530)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold",
       bg="grey", command=Exit).place(x=1000, y=610)

root.mainloop()
